from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_SRC = Path(r'C:/Users/AUROS/AppData/Local/hermes/cache/documents/doc_55d0640c0ae4_Copy of Raihan KataUHA.xlsx')


def cell(ws, coord: str) -> Any:
    return ws[coord].value


def n(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace('Rp', '').replace('%', '').replace(',', '').strip()
    if s in {'', '-', '–', '—'}:
        return 0.0
    return float(s or 0)


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or '')[:10]


def pct(part: float, whole: float) -> float:
    return round((part / whole) * 100, 2) if whole else 0.0


def find_value_right_of_label(ws, label_text: str) -> Any:
    wanted = label_text.strip().lower()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip().lower() == wanted:
                return ws.cell(row, col + 1).value
    return None


def find_row_with_label(ws, label_text: str, col: int = 2) -> int | None:
    wanted = label_text.strip().lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip().lower() == wanted:
            return row
    return None


def find_sheet_name(wb, episode: int) -> str:
    exact = f'RAIHAN KU{episode}'
    if exact in wb.sheetnames:
        return exact
    matches = [s for s in wb.sheetnames if re.search(rf'\bKU\s*{episode}\b', s, flags=re.I)]
    if not matches:
        matches = [s for s in wb.sheetnames if f'KU{episode}' in s.upper().replace(' ', '')]
    if not matches:
        raise ValueError(f'Sheet not found for KU{episode}')
    non_reschedule = [s for s in matches if 'reschedule' not in s.lower()]
    return non_reschedule[0] if non_reschedule else matches[0]


def find_value_left_of_label(ws, label_text: str) -> Any:
    wanted = label_text.strip().lower()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip().lower() == wanted:
                return ws.cell(row, col - 1).value if col > 1 else None
    return None


def find_funnel_header_row(ws) -> int:
    row = find_row_with_label(ws, 'Funnel', col=2)
    if not row:
        raise ValueError(f'Funnel header row not found in {ws.title}')
    return row


def header_map(ws, row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    aliases = {
        'funnel': 'name',
        'linkclick': 'clicks',
        'link click': 'clicks',
        'link klik': 'clicks',
        'addpaymentinfo': 'checkout',
        'purchase': 'paid',
        'transaksi': 'paid',
        'belum infaq': 'unfinished',
        'nominal total transaksi': 'revenue',
    }
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if not isinstance(v, str):
            continue
        key = re.sub(r'\s+', ' ', v.strip().lower())
        compact = key.replace(' ', '')
        canonical = aliases.get(key) or aliases.get(compact)
        if canonical:
            out[canonical] = col
    required = {'name', 'clicks', 'checkout', 'paid', 'revenue'}
    missing = required - set(out)
    if missing:
        raise ValueError(f'Missing funnel columns in {ws.title}: {sorted(missing)} from row {row}: {out}')
    return out


def row_value(ws, row: int, cols: dict[str, int], key: str) -> Any:
    col = cols.get(key)
    return ws.cell(row, col).value if col else None


def parse_daily_revenue(ws) -> list[dict[str, Any]]:
    start_heading = find_row_with_label(ws, 'Monitoring Revenue Harian', col=2)
    if not start_heading:
        return []
    start_row = start_heading + 2  # skip Tanggal/Jumlah header
    history: list[dict[str, Any]] = []
    for row in range(start_row, ws.max_row + 1):
        raw_date = ws.cell(row, 2).value
        amount = n(ws.cell(row, 3).value)
        if isinstance(raw_date, str) and raw_date.strip().lower() == 'grand total':
            break
        if not raw_date and not amount:
            if history:
                break
            continue
        if isinstance(raw_date, (datetime, date)):
            history.append({'date': iso_date(raw_date), 'revenue': int(round(amount))})
    return history


def title_case_episode(label: str, episode: int) -> str:
    return (label or f'KATAUHA {episode}').strip().title().replace('Katauha', 'KataUHA')


def parse_legacy_episode(src: Path, episode: int, sheet_name_override: str | None = None) -> tuple[dict[str, Any], dict[str, Any]]:
    wb = load_workbook(src, data_only=True)
    sheet_name = sheet_name_override or find_sheet_name(wb, episode)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet not found: {sheet_name}')
    ws = wb[sheet_name]

    episode_label = title_case_episode(str(cell(ws, 'B2') or ''), episode)
    event_start = cell(ws, 'B3')
    event_title = str(cell(ws, 'B4') or '').strip()
    last_report_at = cell(ws, 'B5')

    revenue = int(round(n(cell(ws, 'C7'))))
    paid = int(round(n(cell(ws, 'C8'))))
    checkout = int(round(n(cell(ws, 'C9'))))
    unfinished = max(checkout - paid, 0)

    target_revenue = int(round(n(cell(ws, 'D7'))))
    target_paid = int(round(n(cell(ws, 'D8'))))
    target_checkout = int(round(n(cell(ws, 'D9'))))
    target_avg_infaq = int(round(n(cell(ws, 'D10'))))
    target_checkout_to_paid_pct = n(cell(ws, 'D11')) * 100 if n(cell(ws, 'D11')) <= 1 else n(cell(ws, 'D11'))

    meta_spend = n(cell(ws, 'G10'))
    meta_conversion = n(cell(ws, 'G11'))
    ppn = n(find_value_left_of_label(ws, 'PPN 11%')) or n(find_value_left_of_label(ws, 'PPN 12%'))
    if str(cell(ws, 'H13') or '').strip().lower() == 'roas':
        sheet_roas = n(cell(ws, 'G13'))
    elif str(cell(ws, 'H12') or '').strip().lower() == 'roas':
        sheet_roas = n(cell(ws, 'G12'))
    elif str(cell(ws, 'H8') or '').strip().lower() == 'roas':
        sheet_roas = n(cell(ws, 'G8'))
    else:
        sheet_roas = 0.0
    spend_with_ppn = n(find_value_left_of_label(ws, 'Total Ads Spent')) or (meta_spend + ppn)

    promo_end = iso_date(find_value_right_of_label(ws, 'Akhir Promosi'))
    promo_start = iso_date(find_value_right_of_label(ws, 'Start Promosi') or find_value_right_of_label(ws, 'Start'))
    promo_days = int(round(n(find_value_right_of_label(ws, 'Total Waktu Promosi')) or n(find_value_right_of_label(ws, 'Promosi Berjalan')) or 0))

    header_row = find_funnel_header_row(ws)
    cols = header_map(ws, header_row)
    total_row = find_row_with_label(ws, 'TOTAL', col=2)
    if not total_row:
        raise ValueError(f'TOTAL row not found in {sheet_name}')

    funnel_rows = []
    for row in range(header_row + 1, total_row):
        name = row_value(ws, row, cols, 'name')
        if not name:
            continue
        product_name = str(name).strip()
        link_clicks = int(round(n(row_value(ws, row, cols, 'clicks'))))
        row_checkout = int(round(n(row_value(ws, row, cols, 'checkout'))))
        row_paid = int(round(n(row_value(ws, row, cols, 'paid'))))
        row_unfinished = int(round(n(row_value(ws, row, cols, 'unfinished')))) if 'unfinished' in cols else max(row_checkout - row_paid, 0)
        row_revenue = int(round(n(row_value(ws, row, cols, 'revenue'))))
        if not any([link_clicks, row_checkout, row_paid, row_unfinished, row_revenue]):
            continue
        suffix = re.sub(rf'#?\s*katauha\s*{episode}', '', product_name, flags=re.I).strip(' -').lower() or 'main'
        click_source = 'Meta/TikTok Ads' if re.search(r'ads|meta|tiktok', product_name, flags=re.I) else 'legacy manual'
        funnel_rows.append({
            'suffix': suffix,
            'productName': product_name,
            'productId': '',
            'viewsYear': 0,
            'viewsMonth': 0,
            'checkout': row_checkout,
            'paid': row_paid,
            'unfinished': row_unfinished,
            'revenue': row_revenue,
            'avgInfaq': round(row_revenue / row_paid, 2) if row_paid else 0,
            'checkoutToPaidPct': pct(row_paid, row_checkout),
            'viewToCheckoutPct': 0,
            'viewToPaidPct': 0,
            'firstTransactionDate': '',
            'sidLinks': '',
            'linkClicks': link_clicks,
            'uniqueClicks': 0,
            'clickSource': click_source,
            'clickToCheckoutPct': pct(row_checkout, link_clicks),
            'clickToPaidPct': pct(row_paid, link_clicks),
            'recommendation': 'Legacy import - cek detail source jika perlu audit lanjutan',
        })

    # Some legacy sheets have hidden/manual total deltas that are not represented by a named funnel row.
    # Add an explicit adjustment row so dashboard aggregates reconcile while making the caveat visible.
    delta_revenue = revenue - sum(r['revenue'] for r in funnel_rows)
    delta_paid = paid - sum(r['paid'] for r in funnel_rows)
    delta_checkout = checkout - sum(r['checkout'] for r in funnel_rows)
    delta_clicks = int(round(n(row_value(ws, total_row, cols, 'clicks')))) - sum(r['linkClicks'] for r in funnel_rows)
    if any([delta_revenue, delta_paid, delta_checkout]):
        adj_unfinished = max(delta_checkout - delta_paid, 0)
        funnel_rows.append({
            'suffix': 'legacy-adjustment',
            'productName': 'Legacy Adjustment - Sheet Total Delta',
            'productId': '',
            'viewsYear': 0,
            'viewsMonth': 0,
            'checkout': delta_checkout,
            'paid': delta_paid,
            'unfinished': adj_unfinished,
            'revenue': delta_revenue,
            'avgInfaq': round(delta_revenue / delta_paid, 2) if delta_paid else 0,
            'checkoutToPaidPct': pct(delta_paid, delta_checkout),
            'viewToCheckoutPct': 0,
            'viewToPaidPct': 0,
            'firstTransactionDate': '',
            'sidLinks': '',
            'linkClicks': delta_clicks,
            'uniqueClicks': 0,
            'clickSource': 'legacy adjustment',
            'clickToCheckoutPct': pct(delta_checkout, delta_clicks),
            'clickToPaidPct': pct(delta_paid, delta_clicks),
            'recommendation': 'Legacy adjustment row added so named funnel rows reconcile to sheet summary totals.',
        })

    row_link_clicks = sum(r['linkClicks'] for r in funnel_rows)
    summary_clicks = int(round(n(row_value(ws, total_row, cols, 'clicks'))))
    sid_total_clicks = summary_clicks or row_link_clicks
    ads_funnel_revenue = int(round(meta_conversion))
    roas_ads_funnel_with_ppn = round(ads_funnel_revenue / spend_with_ppn, 4) if spend_with_ppn else 0

    daily_revenue_history = parse_daily_revenue(ws)
    daily_fallback = False
    if not daily_revenue_history:
        daily_fallback = True
        daily_revenue_history = [{'date': promo_end or iso_date(last_report_at), 'revenue': revenue}]
    daily_revenue_sum = sum(r['revenue'] for r in daily_revenue_history)

    warnings = []
    if any([delta_revenue, delta_paid, delta_checkout]):
        warnings.append(
            f'Legacy sheet summary totals differed from named funnel rows; added explicit adjustment row with revenue delta {delta_revenue}, paid delta {delta_paid}, checkout delta {delta_checkout}.'
        )
    if daily_fallback:
        warnings.append(f'Monitoring revenue harian tidak tersedia lengkap di sheet {sheet_name}; dashboard memakai satu titik cumulative revenue.')
    else:
        warnings.append('Monitoring revenue harian tersedia dari sheet legacy; angka dipakai untuk chart, tetapi format legacy tetap partial dibanding dashboard baru.')
    warnings.extend([
        'Unique clicks, product IDs, and live source URLs tidak tersedia di format legacy.',
        'Meta/campaign breakdown tidak tersedia; memakai summary spend/conversion/ROAS dari sheet.',
    ])
    if daily_revenue_sum and round(daily_revenue_sum) != round(revenue):
        warnings.append(
            f'Total monitoring revenue harian ({daily_revenue_sum}) berbeda dari summary/funnel revenue ({revenue}); dashboard mempertahankan summary/funnel sebagai KPI utama.'
        )
    if summary_clicks and round(row_link_clicks) != round(summary_clicks):
        warnings.append(
            f'Total link click detail funnel ({row_link_clicks}) berbeda dari summary sheet ({summary_clicks}); KPI tracked clicks memakai summary sheet.'
        )

    out = {
        'episode': episode,
        'episodeLabel': f'KataUHA {episode}',
        'eventTitle': event_title,
        'eventStart': iso_date(event_start),
        'promo_start': promo_start,
        'promo_end': promo_end,
        'promo_days': promo_days,
        'aggregate': {
            'viewsYear': 0,
            'viewsMonth': 0,
            'checkout': checkout,
            'paid': paid,
            'unfinished': unfinished,
            'revenue': revenue,
        },
        'kpi_targets': {
            'targetRevenue': target_revenue,
            'targetPaid': target_paid,
            'targetCheckout': target_checkout,
            'targetAvgInfaq': target_avg_infaq,
            'targetCheckoutToPaidPct': round(target_checkout_to_paid_pct, 2),
        },
        'daily_revenue_history': daily_revenue_history,
        'sid_total_clicks': sid_total_clicks,
        'sid_total_unique': 0,
        'meta_summary': {
            'spend': meta_spend,
            'ppn': ppn,
            'spendWithPpn': spend_with_ppn,
            'metaRevenue': meta_conversion,
            'metaRoas': sheet_roas,
            'adsFunnelRevenue': ads_funnel_revenue,
            'roasAdsFunnelWithPpn': roas_ads_funnel_with_ppn,
            'campaignCount': 0,
        },
        'funnel_rows': sorted(funnel_rows, key=lambda r: r['revenue'], reverse=True),
        'legacy_source': {
            'workbook': str(src),
            'sheet': sheet_name,
            'episodeLabelFromSheet': episode_label,
            'lastReportAt': last_report_at.isoformat() if isinstance(last_report_at, datetime) else str(last_report_at),
            'dataQuality': 'legacy-normalized-partial',
            'warnings': warnings,
        },
    }

    audit = {
        'ok': True,
        'episode': episode,
        'sourceSheet': sheet_name,
        'checks': {
            'summaryRevenue': revenue,
            'sumFunnelRevenue': sum(r['revenue'] for r in funnel_rows),
            'summaryPaid': paid,
            'sumFunnelPaid': sum(r['paid'] for r in funnel_rows),
            'summaryCheckout': checkout,
            'sumFunnelCheckout': sum(r['checkout'] for r in funnel_rows),
            'summaryClicks': summary_clicks,
            'sumFunnelClicks': row_link_clicks,
            'adsSpend': meta_spend,
            'adsPpn': ppn,
            'adsSpendWithPpn': spend_with_ppn,
            'adsFunnelRevenue': ads_funnel_revenue,
            'sheetRoasNoPpn': sheet_roas,
            'computedRoasWithPpn': roas_ads_funnel_with_ppn,
            'dailyRevenuePoints': len(daily_revenue_history),
            'dailyRevenueSum': daily_revenue_sum,
            'dailyRevenueVsSummaryDiff': daily_revenue_sum - revenue,
        },
        'warnings': warnings,
    }
    for a, b in [('summaryRevenue', 'sumFunnelRevenue'), ('summaryPaid', 'sumFunnelPaid'), ('summaryCheckout', 'sumFunnelCheckout')]:
        if round(audit['checks'][a]) != round(audit['checks'][b]):
            audit['ok'] = False
            audit.setdefault('mismatches', []).append({'left': a, 'right': b, 'values': [audit['checks'][a], audit['checks'][b]]})

    return out, audit


def combine_ku25_reschedule(src: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    first, first_audit = parse_legacy_episode(src, 25, 'RAIHAN KU25')
    resched, resched_audit = parse_legacy_episode(src, 25, 'RAIHAN KU25-Reschedule')

    combined = dict(resched)
    combined['episodeLabel'] = 'KataUHA 25'
    combined['eventTitle'] = resched.get('eventTitle') or first.get('eventTitle')
    combined['eventStart'] = resched.get('eventStart') or first.get('eventStart')
    combined['promo_start'] = first.get('promo_start') or resched.get('promo_start')
    combined['promo_end'] = resched.get('promo_end') or resched.get('eventStart')
    combined['promo_days'] = max(int(first.get('promo_days') or 0) + int(resched.get('promo_days') or 0), int(resched.get('promo_days') or 0))

    for key in ['revenue', 'paid', 'checkout', 'unfinished']:
        combined['aggregate'][key] = (first.get('aggregate') or {}).get(key, 0) + (resched.get('aggregate') or {}).get(key, 0)

    combined['sid_total_clicks'] = int(first.get('sid_total_clicks') or 0) + int(resched.get('sid_total_clicks') or 0)
    combined['sid_total_unique'] = 0
    combined['funnel_rows'] = []
    for phase, data in [('hold', first), ('reschedule', resched)]:
        for row in data.get('funnel_rows') or []:
            rr = dict(row)
            rr['suffix'] = f"{phase}-{rr.get('suffix') or 'main'}"
            rr['productName'] = f"{phase.title()} - {rr.get('productName') or ''}".strip()
            combined['funnel_rows'].append(rr)
    combined['funnel_rows'] = sorted(combined['funnel_rows'], key=lambda r: r.get('revenue') or 0, reverse=True)

    fm = first.get('meta_summary') or {}
    rm = resched.get('meta_summary') or {}
    spend = float(fm.get('spend') or 0) + float(rm.get('spend') or 0)
    ppn = float(fm.get('ppn') or 0) + float(rm.get('ppn') or 0)
    spend_with_ppn = float(fm.get('spendWithPpn') or 0) + float(rm.get('spendWithPpn') or 0)
    ads_rev = float(fm.get('adsFunnelRevenue') or 0) + float(rm.get('adsFunnelRevenue') or 0)
    combined['meta_summary'] = {
        'spend': spend,
        'ppn': ppn,
        'spendWithPpn': spend_with_ppn,
        'metaRevenue': ads_rev,
        'metaRoas': round(ads_rev / spend, 4) if spend else 0,
        'adsFunnelRevenue': ads_rev,
        'roasAdsFunnelWithPpn': round(ads_rev / spend_with_ppn, 4) if spend_with_ppn else 0,
        'campaignCount': 0,
    }

    # User confirmed this reschedule sheet's Monitoring Revenue Harian is the combined monitoring source.
    combined['daily_revenue_history'] = resched.get('daily_revenue_history') or []
    daily_sum = sum(int(round(float(r.get('revenue') or 0))) for r in combined['daily_revenue_history'])
    aggregate = combined['aggregate']
    funnel_rows = combined['funnel_rows']
    combined['legacy_source'] = {
        'workbook': str(src),
        'sheet': 'RAIHAN KU25 + RAIHAN KU25-Reschedule',
        'episodeLabelFromSheet': 'KATAUHA 25 + KATAUHA 25 (Reschedule)',
        'lastReportAt': (resched.get('legacy_source') or {}).get('lastReportAt'),
        'dataQuality': 'legacy-normalized-partial-combined-reschedule',
        'warnings': [
            'KU25 terdiri dari dua fase promosi: RAIHAN KU25 (hold) dan RAIHAN KU25-Reschedule; KPI agregat menjumlahkan keduanya sebagai satu episode.',
            'Monitoring Revenue Harian memakai sheet RAIHAN KU25-Reschedule sesuai konfirmasi user karena sudah gabungan kedua fase promosi.',
            'Unique clicks, product IDs, and live source URLs tidak tersedia di format legacy.',
            'Meta/campaign breakdown tidak tersedia; memakai summary spend/conversion/ROAS dari kedua sheet.',
        ],
    }
    if daily_sum and round(daily_sum) != round(aggregate['revenue']):
        combined['legacy_source']['warnings'].append(
            f'Total monitoring revenue harian ({daily_sum}) berbeda dari combined summary/funnel revenue ({aggregate["revenue"]}); dashboard mempertahankan combined summary/funnel sebagai KPI utama.'
        )

    audit = {
        'ok': True,
        'episode': 25,
        'sourceSheet': 'RAIHAN KU25 + RAIHAN KU25-Reschedule',
        'checks': {
            'summaryRevenue': aggregate['revenue'],
            'sumFunnelRevenue': sum(r.get('revenue') or 0 for r in funnel_rows),
            'summaryPaid': aggregate['paid'],
            'sumFunnelPaid': sum(r.get('paid') or 0 for r in funnel_rows),
            'summaryCheckout': aggregate['checkout'],
            'sumFunnelCheckout': sum(r.get('checkout') or 0 for r in funnel_rows),
            'summaryClicks': combined['sid_total_clicks'],
            'sumFunnelClicks': sum(r.get('linkClicks') or 0 for r in funnel_rows),
            'adsSpend': spend,
            'adsPpn': ppn,
            'adsSpendWithPpn': spend_with_ppn,
            'adsFunnelRevenue': ads_rev,
            'sheetRoasNoPpn': combined['meta_summary']['metaRoas'],
            'computedRoasWithPpn': combined['meta_summary']['roasAdsFunnelWithPpn'],
            'dailyRevenuePoints': len(combined['daily_revenue_history']),
            'dailyRevenueSum': daily_sum,
            'dailyRevenueVsSummaryDiff': daily_sum - aggregate['revenue'],
            'firstPhaseAuditOk': first_audit.get('ok'),
            'rescheduleAuditOk': resched_audit.get('ok'),
        },
        'warnings': combined['legacy_source']['warnings'],
    }
    for a, b in [('summaryRevenue', 'sumFunnelRevenue'), ('summaryPaid', 'sumFunnelPaid'), ('summaryCheckout', 'sumFunnelCheckout')]:
        if round(audit['checks'][a]) != round(audit['checks'][b]):
            audit['ok'] = False
            audit.setdefault('mismatches', []).append({'left': a, 'right': b, 'values': [audit['checks'][a], audit['checks'][b]]})
    return combined, audit


def write_episode(src: Path, episode: int) -> dict[str, Any]:
    out, audit = combine_ku25_reschedule(src) if episode == 25 else parse_legacy_episode(src, episode)
    out_dir = Path(f'C:/Users/AUROS/reports/katauha{episode}')
    out_dir.mkdir(parents=True, exist_ok=True)
    out_json = out_dir / f'katauha{episode}_dashboard_updated.json'
    audit_json = out_dir / f'katauha{episode}_legacy_import_audit.json'
    out_json.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
    audit_json.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    return {'ok': audit['ok'], 'json': str(out_json), 'audit': str(audit_json), **audit['checks'], 'warnings': audit['warnings']}


def main() -> int:
    parser = argparse.ArgumentParser(description='Import Raihan KataUHA legacy workbook episodes into normalized dashboard JSON.')
    parser.add_argument('episodes', nargs='+', type=int, help='Episode numbers, e.g. 35 36 37')
    parser.add_argument('--src', type=Path, default=DEFAULT_SRC)
    args = parser.parse_args()

    results = [write_episode(args.src, ep) for ep in args.episodes]
    print(json.dumps({'ok': all(r['ok'] for r in results), 'results': results}, ensure_ascii=False, indent=2))
    return 0 if all(r['ok'] for r in results) else 1


if __name__ == '__main__':
    raise SystemExit(main())
