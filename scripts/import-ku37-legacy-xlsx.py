from __future__ import annotations

import json
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC = Path(r'C:/Users/AUROS/AppData/Local/hermes/cache/documents/doc_ab2e62143294_Copy of Raihan KataUHA.xlsx')
SHEET = 'RAIHAN KU37'
OUT_DIR = Path(r'C:/Users/AUROS/reports/katauha37')
OUT_JSON = OUT_DIR / 'katauha37_dashboard_updated.json'
AUDIT_JSON = OUT_DIR / 'katauha37_legacy_import_audit.json'


def cell(ws, coord: str) -> Any:
    return ws[coord].value


def n(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace('Rp', '').replace('%', '').replace(',', '').strip()
    return float(s or 0)


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or '')[:10]


def pct(part: float, whole: float) -> float:
    return round((part / whole) * 100, 2) if whole else 0.0


def parse_daily_revenue(ws) -> list[dict[str, Any]]:
    """Parse the legacy Monitoring Revenue Harian table.

    KU37 stores daily values in B:C under the "Monitoring Revenue Harian"
    heading. Values are already daily amounts, not cumulative values.
    """
    start_row = None
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 2).value
        if isinstance(label, str) and label.strip().lower() == 'monitoring revenue harian':
            start_row = row + 2  # skip header row "Tanggal" / "Jumlah"
            break
    if not start_row:
        return []

    history: list[dict[str, Any]] = []
    for row in range(start_row, ws.max_row + 1):
        raw_date = ws.cell(row, 2).value
        amount = n(ws.cell(row, 3).value)
        if isinstance(raw_date, str) and raw_date.strip().lower() == 'grand total':
            break
        if not raw_date and not amount:
            if history:
                break
            continue
        if isinstance(raw_date, (datetime, date)):
            history.append({'date': iso_date(raw_date), 'revenue': int(round(amount))})
    return history


wb = load_workbook(SRC, data_only=True)
ws = wb[SHEET]

# Summary block B:G.
episode_label = str(cell(ws, 'B2') or 'KATAUHA 37').strip().title().replace('Katauha', 'KataUHA')
event_start = cell(ws, 'B3')
event_title = str(cell(ws, 'B4') or '').strip()
last_report_at = cell(ws, 'B5')

revenue = int(round(n(cell(ws, 'C7'))))
paid = int(round(n(cell(ws, 'C8'))))
checkout = int(round(n(cell(ws, 'C9'))))
avg_infaq = n(cell(ws, 'C10'))
checkout_to_paid = n(cell(ws, 'C11')) * 100 if n(cell(ws, 'C11')) <= 1 else n(cell(ws, 'C11'))
unfinished = max(checkout - paid, 0)

target_revenue = int(round(n(cell(ws, 'D7'))))
target_paid = int(round(n(cell(ws, 'D8'))))
target_checkout = int(round(n(cell(ws, 'D9'))))
target_avg_infaq = int(round(n(cell(ws, 'D10'))))
target_checkout_to_paid_pct = n(cell(ws, 'D11')) * 100 if n(cell(ws, 'D11')) <= 1 else n(cell(ws, 'D11'))

meta_spend = n(cell(ws, 'G10'))
meta_conversion = n(cell(ws, 'G11'))
ppn = n(cell(ws, 'G12'))
sheet_roas = n(cell(ws, 'G13'))
spend_with_ppn = n(cell(ws, 'G15')) or (meta_spend + ppn)

promo_end = iso_date(cell(ws, 'P37'))
promo_start = iso_date(cell(ws, 'P38'))
promo_days = int(round(n(cell(ws, 'P43')) or 0))
raihan_hari_ini = n(cell(ws, 'P51'))

# Funnel rows B:M, rows 18-33; row 34 is TOTAL.
# Revised KU37 has row 43 "LP Test After Launch" as an explanatory detail;
# the corrected amount is already reflected in row 23 and row 34 totals, so
# row 43 is intentionally not imported to avoid double counting.
funnel_rows = []
for row in range(18, 34):
    name = cell(ws, f'B{row}')
    if not name:
        continue
    product_name = str(name).strip()
    link_clicks = int(round(n(cell(ws, f'C{row}'))))
    row_checkout = int(round(n(cell(ws, f'G{row}'))))
    row_paid = int(round(n(cell(ws, f'H{row}'))))
    row_unfinished = int(round(n(cell(ws, f'I{row}'))))
    row_revenue = int(round(n(cell(ws, f'J{row}'))))
    if not any([link_clicks, row_checkout, row_paid, row_unfinished, row_revenue]):
        continue
    suffix = product_name.lower().replace('#katauha 37', '').replace('katauha 37', '').strip(' -') or 'main'
    click_source = 'Meta/TikTok Ads' if 'ads' in product_name.lower() or 'meta' in product_name.lower() else 'legacy manual'
    funnel_rows.append({
        'suffix': suffix,
        'productName': product_name,
        'productId': '',
        'viewsYear': 0,
        'viewsMonth': 0,
        'checkout': row_checkout,
        'paid': row_paid,
        'unfinished': row_unfinished,
        'revenue': row_revenue,
        'avgInfaq': round(row_revenue / row_paid, 2) if row_paid else 0,
        'checkoutToPaidPct': pct(row_paid, row_checkout),
        'viewToCheckoutPct': 0,
        'viewToPaidPct': 0,
        'firstTransactionDate': '',
        'sidLinks': '',
        'linkClicks': link_clicks,
        'uniqueClicks': 0,
        'clickSource': click_source,
        'clickToCheckoutPct': pct(row_checkout, link_clicks),
        'clickToPaidPct': pct(row_paid, link_clicks),
        'recommendation': 'Legacy import - cek detail source jika perlu audit lanjutan'
    })

sid_total_clicks = sum(r['linkClicks'] for r in funnel_rows)
ads_funnel_revenue = int(round(meta_conversion))
roas_ads_funnel_with_ppn = round(ads_funnel_revenue / spend_with_ppn, 4) if spend_with_ppn else 0

daily_revenue_history = parse_daily_revenue(ws)
if not daily_revenue_history:
    daily_revenue_history = [
        {'date': promo_end or iso_date(last_report_at), 'revenue': revenue}
    ]
daily_revenue_sum = sum(r['revenue'] for r in daily_revenue_history)

out = {
    'episode': 37,
    'episodeLabel': 'KataUHA 37',
    'eventTitle': event_title,
    'eventStart': iso_date(event_start),
    'promo_start': promo_start,
    'promo_end': promo_end,
    'promo_days': promo_days,
    'aggregate': {
        'viewsYear': 0,
        'viewsMonth': 0,
        'checkout': checkout,
        'paid': paid,
        'unfinished': unfinished,
        'revenue': revenue,
    },
    'kpi_targets': {
        'targetRevenue': target_revenue,
        'targetPaid': target_paid,
        'targetCheckout': target_checkout,
        'targetAvgInfaq': target_avg_infaq,
        'targetCheckoutToPaidPct': round(target_checkout_to_paid_pct, 2),
    },
    'daily_revenue_history': daily_revenue_history,
    'sid_total_clicks': sid_total_clicks,
    'sid_total_unique': 0,
    'meta_summary': {
        'spend': meta_spend,
        'ppn': ppn,
        'spendWithPpn': spend_with_ppn,
        'metaRevenue': meta_conversion,
        'metaRoas': sheet_roas,
        'adsFunnelRevenue': ads_funnel_revenue,
        'roasAdsFunnelWithPpn': roas_ads_funnel_with_ppn,
        'campaignCount': 0,
    },
    'funnel_rows': sorted(funnel_rows, key=lambda r: r['revenue'], reverse=True),
    'legacy_source': {
        'workbook': str(SRC),
        'sheet': SHEET,
        'lastReportAt': last_report_at.isoformat() if isinstance(last_report_at, datetime) else str(last_report_at),
        'dataQuality': 'legacy-normalized-partial',
        'warnings': [
            'Monitoring revenue harian tersedia dari sheet legacy; angka dipakai untuk chart, tetapi format legacy tetap partial dibanding dashboard baru.',
            'Unique clicks, product IDs, and live source URLs tidak tersedia di format legacy.',
            'Meta/campaign breakdown tidak tersedia; memakai summary spend/conversion/ROAS dari sheet.'
        ],
    }
}

# Validation against sheet totals.
audit = {
    'ok': True,
    'episode': 37,
    'sourceSheet': SHEET,
    'checks': {
        'summaryRevenue': revenue,
        'sumFunnelRevenue': sum(r['revenue'] for r in funnel_rows),
        'summaryPaid': paid,
        'sumFunnelPaid': sum(r['paid'] for r in funnel_rows),
        'summaryCheckout': checkout,
        'sumFunnelCheckout': sum(r['checkout'] for r in funnel_rows),
        'summaryClicks': int(round(n(cell(ws, 'C34')))),
        'sumFunnelClicks': sid_total_clicks,
        'adsSpend': meta_spend,
        'adsPpn': ppn,
        'adsSpendWithPpn': spend_with_ppn,
        'adsFunnelRevenue': ads_funnel_revenue,
        'sheetRoasNoPpn': sheet_roas,
        'computedRoasWithPpn': roas_ads_funnel_with_ppn,
        'dailyRevenuePoints': len(daily_revenue_history),
        'dailyRevenueSum': daily_revenue_sum,
        'dailyRevenueVsSummaryDiff': daily_revenue_sum - revenue,
    },
    'warnings': out['legacy_source']['warnings'],
}
for a, b in [('summaryRevenue', 'sumFunnelRevenue'), ('summaryPaid', 'sumFunnelPaid'), ('summaryCheckout', 'sumFunnelCheckout'), ('summaryClicks', 'sumFunnelClicks')]:
    if round(audit['checks'][a]) != round(audit['checks'][b]):
        audit['ok'] = False
        audit.setdefault('mismatches', []).append({'left': a, 'right': b, 'values': [audit['checks'][a], audit['checks'][b]]})
if daily_revenue_sum and round(daily_revenue_sum) != round(revenue):
    audit['warnings'].append(
        f"Total monitoring revenue harian ({daily_revenue_sum}) berbeda dari summary/funnel revenue ({revenue}); dashboard mempertahankan summary/funnel sebagai KPI utama."
    )

OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
AUDIT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps({'ok': audit['ok'], 'json': str(OUT_JSON), 'audit': str(AUDIT_JSON), **audit['checks'], 'warnings': audit['warnings']}, ensure_ascii=False, indent=2))
raise SystemExit(0 if audit['ok'] else 1)
